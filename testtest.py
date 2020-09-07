# logistics_dic = {
#     '고양1': {
#         '발주건수': 3,
#         '발주수량': [8, 22, 20],
#         '발주아이템': ['말랑하니 오가닉 속싸개_엘라스틴_그레이도트M', '말랑하니 오가닉속싸개_엘라스틴_캐럿핑크S', '말랑하니 오가닉속싸개_엘라스틴_캐럿핑크M'],
#         '발주번호': [19351878, 19351878, 19346395],
#     }
# }

# print(logistics_dic.get(''))


# print(logistics_dic.get('고양2', 'NOTHING'))
# print(logistics_dic.get('고양1').get('발주번호')[2])
# logistics_dic = {}
# for i in range(city_numbers):
#     logistics_dic[city_series[i]] = df[cities == city_series[i]].shape[0]

# is_goyang1 = df[df['물류센터'] == '덕평1']
# print(is_goyang1.shape[0])
# # print(type(is_goyang1.shape[1]))
# for i in range(is_goyang1.shape[1]):
#     order_number.append(is_goyang1.iloc[i].loc['발주번호'])
#     order_quantity.append(is_goyang1.iloc[i].loc['발주수량'])
#     order_item.append(is_goyang1.iloc[i].loc['SKU 이름'])
# print("덕평1 발주번호: ", order_number)
# print("덕평1 발주수량: ", order_quantity)
# print("덕평1 발주아이템: ", order_item)

import datetime
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import pandas as pd
import city_info as ci


df = pd.read_csv(
    '/Users/martinkim/GITHUB/00_Automated System/02_Logistics-team/logistics_tracking-number/PO_SKU_LIST_20200904150707.csv')
order_count = df.shape[0]
cities = df['물류센터']
city_series = cities.drop_duplicates()
city_series.reset_index(drop=True, inplace=True)
city_numbers = len(city_series)
order_info = [[0 for col in range(2)] for row in range(city_numbers)]
for i in range(city_numbers):
    order_info[i][0] = city_series[i]
    order_info[i][1] = df[cities == city_series[i]].shape[0]

print(
    f'총 발주 물량은 {order_count}개이며, 총 {city_numbers}군데의 도시에서 {order_info}이다')

# CREATE EXCEL
today_date = datetime.datetime.now().strftime('%Y%m%d')

wb = openpyxl.Workbook()
ws = wb.active
# Initialize
ws.title = today_date + '입고예정'
ws['A1'] = '이름'
ws['B1'] = '전화1'
ws['C1'] = '전화2'
ws['D1'] = '우편번호'
ws['E1'] = '주소'
ws['F1'] = '수량'
ws['G1'] = '품목'
ws['H1'] = '배송시요구사항'
ws['I1'] = '사이트'
ws['J1'] = '순번'
ws['K1'] = '주문번호'
ws['L1'] = '운송장번호'
ws['M1'] = '받는분'
# Make cells Bold / Aligned / Bg-colored
cell_list = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1',
             'G1', 'H1', 'I1', 'J1', 'K1', 'L1', 'M1']
for i in range(len(cell_list)):
    ws[cell_list[i]].font = Font(bold=True, color='FFFFFF')
    ws[cell_list[i]].fill = PatternFill(
        fgColor='8B0000', patternType='solid')
    ws[cell_list[i]].alignment = Alignment(
        horizontal='center', vertical='center')
# Differentiate two columns for future input
ws['L1'].fill = PatternFill(fgColor='ADD8E6', patternType='solid')
ws['M1'].fill = PatternFill(fgColor='ADD8E6', patternType='solid')

# INPUT VALUE INTO EXCEL
order_quantity = []
order_number = []
order_item = []
start = 2

for i in range(city_numbers):
    city = df[df['물류센터'] == order_info[i][0]]

    for j in range(order_info[i][1]):
        order_quantity.append(city.iloc[j].loc['발주수량'])
        order_item.append(city.iloc[j].loc['SKU 이름'])
        order_number.append(city.iloc[j].loc['발주번호'])

    print(f'{order_info[i][0]} 발주수량: {order_quantity}')
    print(f'{order_info[i][0]} 발주아이템: {order_item}')
    print(f'{order_info[i][0]} 발주번호: {order_number}')

    for k in range(order_info[i][1]):
        if ws['A'+str(start)].value != None:
            while (ws['A'+str(start)].value != None):
                start += 1
        ws['A'+str(start)] = order_info[i][0]
        ws['F'+str(start)] = order_quantity[k]
        ws['G'+str(start)] = order_item[k]
        ws['H'+str(start)] = "발주번호: " + str(order_number[k])
        ws['I'+str(start)] = '쿠팡'

        if ws['A'+str(start)].value == '고양1':
            ws['B'+str(start)] = ci.GOYANG1.get('phone1')
            ws['C'+str(start)] = ci.GOYANG1.get('phone2')
            ws['D'+str(start)] = ci.GOYANG1.get('zip')
            ws['E'+str(start)] = ci.GOYANG1.get('address')
        elif ws['A'+str(start)].value == '광주':
            ws['B'+str(start)] = ci.GWANGJU.get('phone1')
            ws['C'+str(start)] = ci.GWANGJU.get('phone2')
            ws['D'+str(start)] = ci.GWANGJU.get('zip')
            ws['E'+str(start)] = ci.GWANGJU.get('address')
        elif ws['A'+str(start)].value == '대구2':
            ws['B'+str(start)] = ci.DAEGU2.get('phone1')
            ws['C'+str(start)] = ci.DAEGU2.get('phone2')
            ws['D'+str(start)] = ci.DAEGU2.get('zip')
            ws['E'+str(start)] = ci.DAEGU2.get('address')
        elif ws['A'+str(start)].value == '덕평1':
            ws['B'+str(start)] = ci.DEOKPYEONG1.get('phone1')
            ws['C'+str(start)] = ci.DEOKPYEONG1.get('phone2')
            ws['D'+str(start)] = ci.DEOKPYEONG1.get('zip')
            ws['E'+str(start)] = ci.DEOKPYEONG1.get('address')
        elif ws['A'+str(start)].value == '동탄1':
            ws['B'+str(start)] = ci.DONGTAN1.get('phone1')
            ws['C'+str(start)] = ci.DONGTAN1.get('phone2')
            ws['D'+str(start)] = ci.DONGTAN1.get('zip')
            ws['E'+str(start)] = ci.DONGTAN1.get('address')
        elif ws['A'+str(start)].value == '마장1':
            ws['B'+str(start)] = ci.MAJANG.get('phone1')
            ws['C'+str(start)] = ci.MAJANG.get('phone2')
            ws['D'+str(start)] = ci.MAJANG.get('zip')
            ws['E'+str(start)] = ci.MAJANG.get('address')
        elif ws['A'+str(start)].value == '목천1':
            ws['B'+str(start)] = ci.MOKCHEON1.get('phone1')
            ws['C'+str(start)] = ci.MOKCHEON1.get('phone2')
            ws['D'+str(start)] = ci.MOKCHEON1.get('zip')
            ws['E'+str(start)] = ci.MOKCHEON1.get('address')
        elif ws['A'+str(start)].value == '부천1':
            ws['B'+str(start)] = ci.BUCHEON1.get('phone1')
            ws['C'+str(start)] = ci.BUCHEON1.get('phone2')
            ws['D'+str(start)] = ci.BUCHEON1.get('zip')
            ws['E'+str(start)] = ci.BUCHEON1.get('address')
        elif ws['A'+str(start)].value == '서울':
            ws['B'+str(start)] = ci.SEOUL.get('phone1')
            ws['C'+str(start)] = ci.SEOUL.get('phone2')
            ws['D'+str(start)] = ci.SEOUL.get('zip')
            ws['E'+str(start)] = ci.SEOUL.get('address')
        elif ws['A'+str(start)].value == '안성4':
            ws['B'+str(start)] = ci.ANSEONG4.get('phone1')
            ws['C'+str(start)] = ci.ANSEONG4.get('phone2')
            ws['D'+str(start)] = ci.ANSEONG4.get('zip')
            ws['E'+str(start)] = ci.ANSEONG4.get('address')
        elif ws['A'+str(start)].value == '양산1':
            ws['B'+str(start)] = ci.YANGSAN1.get('phone1')
            ws['C'+str(start)] = ci.YANGSAN1.get('phone2')
            ws['D'+str(start)] = ci.YANGSAN1.get('zip')
            ws['E'+str(start)] = ci.YANGSAN1.get('address')
        elif ws['A'+str(start)].value == '인천1':
            ws['B'+str(start)] = ci.INCHEON1.get('phone1')
            ws['C'+str(start)] = ci.INCHEON1.get('phone2')
            ws['D'+str(start)] = ci.INCHEON1.get('zip')
            ws['E'+str(start)] = ci.INCHEON1.get('address')
        elif ws['A'+str(start)].value == '인천4':
            ws['B'+str(start)] = ci.INCHEON4.get('phone1')
            ws['C'+str(start)] = ci.INCHEON4.get('phone2')
            ws['D'+str(start)] = ci.INCHEON4.get('zip')
            ws['E'+str(start)] = ci.INCHEON4.get('address')
        elif ws['A'+str(start)].value == '인천5':
            ws['B'+str(start)] = ci.INCHEON5.get('phone1')
            ws['C'+str(start)] = ci.INCHEON5.get('phone2')
            ws['D'+str(start)] = ci.INCHEON5.get('zip')
            ws['E'+str(start)] = ci.INCHEON5.get('address')
        elif ws['A'+str(start)].value == '천안':
            ws['B'+str(start)] = ci.CHOENAN.get('phone1')
            ws['C'+str(start)] = ci.CHOENAN.get('phone2')
            ws['D'+str(start)] = ci.CHOENAN.get('zip')
            ws['E'+str(start)] = ci.CHOENAN.get('address')
        elif ws['A'+str(start)].value == '평택1':
            ws['B'+str(start)] = ci.PYOENGTAEK1.get('phone1')
            ws['C'+str(start)] = ci.PYOENGTAEK1.get('phone2')
            ws['D'+str(start)] = ci.PYOENGTAEK1.get('zip')
            ws['E'+str(start)] = ci.PYOENGTAEK1.get('address')
        elif ws['A'+str(start)].value == '호법':
            ws['B'+str(start)] = ci.HOBEOP.get('phone1')
            ws['C'+str(start)] = ci.HOBEOP.get('phone2')
            ws['D'+str(start)] = ci.HOBEOP.get('zip')
            ws['E'+str(start)] = ci.HOBEOP.get('address')

    order_number.clear()
    order_quantity.clear()
    order_item.clear()

wb.save('쿠팡 주소록-정리_' + today_date + '.xlsx')
